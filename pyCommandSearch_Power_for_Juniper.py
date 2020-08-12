# -*- coding: utf-8 -*-

import glob
import openpyxl
import os
import sys
from commonLib import commonLib

class CursorPos:
    def __init__(self, i, j):
        self.i_row = i
        self.i_col = j


def main():
    argv = sys.argv
    argc = len(argv)

    # 引数チェック
    if argc < 4:
        exit_msg(argv[0])

    enable_perfect_match = False
    if argc >= 5:
        if argv[3].upper() == "TRUE":
            enable_perfect_match = True

    enable_excel = False
    if argc >= 4:
        if not os.path.exists(argv[3]):
            print("{0} not found.".format(argv[3]))
            return
        enable_excel = True
    workbook_filenamePath = argv[3]

    target_command = "show chassis power"
    if argc >= 3:
        target_command = argv[2].replace("\"", "")

    target_path = "../log/*.log"
    if argc >= 2:
        target_path = argv[1].replace("\"", "")

    # Excel Workbookを開く
    t_wb = openpyxl.load_workbook(workbook_filenamePath)

    # ファイル一覧の取得
    files_list = glob.glob(target_path)

    """
    for filenamePath in files_list:
        print(filenamePath)
    """

    prompt_char = ["#", "> ", ">"]
    sPos = {}
    for sheetname in t_wb.get_sheet_names():
        sPos[sheetname] = CursorPos(1, 3)

    for filenamePath in files_list:
        # ファイルの内容を読み込む
        f = open(filenamePath, "rt", encoding="ascii")
        contents = f.readlines()
        f.close

        # "show version"コマンドの実行結果を取得する
        contents_target_command, prompt_list = commonLib.get_contents_target_command(contents, "show version", prompt_char, enable_perfect_match)
        print("=================================================================================")
        print(filenamePath)
        for line in contents_target_command:
            if "Model:" in line:
                print(line, end="")
                modelName = line.replace("Model:", "").strip()
                if not sheetExists(t_wb, modelName):
                    t_wb.create_sheet(modelName)
                    sPos[modelName] = CursorPos(2, 3)
                t_ws = t_wb.get_sheet_by_name(modelName)
                t_ws.cell(1, 1).value = "hostname"
                t_ws.cell(1, 2).value = "ipaddr"
                t_ws.cell(1, 3).value = "show chassis power*"
                break

        # output cell
        pos0 = filenamePath.rfind("\\")
        fileName = filenamePath[pos0 + 1:]
        pos1 = fileName.find("_")
        pos2 = fileName.find("_", pos1 + 1)
        pos3 = fileName.find("_", pos2 + 1)

        hostname = fileName[pos1 + 1:pos2]
        ipaddr = fileName[pos2 + 1:pos3]
        t_ws.cell(sPos[t_ws.title].i_row, 1).value = hostname
        t_ws.cell(sPos[t_ws.title].i_row, 2).value = ipaddr
        sPos[t_ws.title].i_col = 3

        # "show chassis power.*"コマンドの実行結果を取得する
        contents_target_command, prompt_list = commonLib.get_contents_target_command(contents, target_command, prompt_char, enable_perfect_match)

        # "show chassis power.*"コマンドの種類を判別する
        powerCommandType = ""
        for line in contents_target_command:
            if "show chassis power detail" in line:
                print("<<Result Type A>>")
                powerCommandType = "show chassis power detail"
                break
            elif "show chassis power-budget-statistics" in line:
                print("<<Result Type B>>")
                powerCommandType = "show chassis power-budget-statistics"
                break
        if powerCommandType == "":
            print("<<Result Type Other>>")
        print("=================================================================================")

        # "show chassis power.*"コマンドの実行結果を表示する
        for line in contents_target_command:
            print(line, end="")

        # "show chassis power.*"コマンドの実行結果のフォーマットを変換して出力する
        if powerCommandType == "show chassis power-budget-statistics":
            # "show chassis power-budget-statistics"の実行結果をテーブル形式のリストに格納する
            fpcs = get_show_chassis_power_budget_statistics(contents_target_command, prompt_list)

            # テーブルへの格納結果を出力する
            print_get_show_chassis_power_budget_statistics(fpcs)

            if enable_excel:
                # テーブルへの格納結果をワークシートに出力する
                output_show_chassis_power_budget_statistics(fpcs, t_ws, sPos)
            #   get_cells(t_ws, 1, 1, 10, 20)

            sPos[t_ws.title].i_row -= 1
            sPos[t_ws.title].i_col = 3

            # "show chassis power detail"の実行結果を１つの文字列としてに格納する
            fpcs = get_show_chassis_power_to_str(contents_target_command, prompt_list)

            # テーブルへの格納結果を出力する
            print_show_chassis_power_detail(fpcs)

            if enable_excel:
                # テーブルへの格納結果をワークシートに出力する
                output_show_chassis_power_detail(fpcs, t_ws, sPos)
            #   get_cells(t_ws, 1, 1, 10, 20)

        elif powerCommandType == "show chassis power detail":
            # "show chassis power detail"の実行結果を１つの文字列としてに格納する
            fpcs = get_show_chassis_power_to_str(contents_target_command, prompt_list)

            # 文字列への格納結果を出力する
            print_show_chassis_power_detail(fpcs)

            if enable_excel:
                # 文字列への格納結果をワークシートに出力する（１セル）
                output_show_chassis_power_detail(fpcs, t_ws, sPos)
            #   get_cells(t_ws, 1, 1, 10, 20)

    # 文字位置を上詰めにする
    worksheets_alignment_set(t_wb)

    res = ""
    while res != "Y" and res != "N":
        res = input("Overwrite (Y/N): ")
        res = res.upper()

    if res == "Y":
        t_wb.save(workbook_filenamePath)


def exit_msg(argv0):
    """使用例を表示する"""
    print("Usage: python {0} <logfiles path> <target command> <target Excel Workbook> <match_flag>\n"
          "\n"
          "example)\n"
          "python {0} \"../log/*.log\" \"show chassis power\" PowerStatus_Juniper.xlsx\n"
          .format(argv0))
    exit(0)


def sheetExists(t_wb, targetname):
    for sheetname in t_wb.sheetnames:
        if targetname == sheetname:
            return True
    return False


def worksheets_alignment_set(t_wb):
    """セルの配置プロパティを変更する"""
    for sheetname in t_wb.get_sheet_names():
        t_ws = t_wb.get_sheet_by_name(sheetname)
        for t_row in t_ws.rows:
            for cell in t_row:
                if cell.column == 3:
                    # 折り返して表示
                    cell.alignment = openpyxl.styles.Alignment(wrapText = True)
                # 上詰めで表示
                cell.alignment = openpyxl.styles.Alignment(vertical = "top")


def get_cells(t_ws, row_start, col_start, row_count, col_count):
    """ 指定セル範囲の内容を取得して表示する"""
    row_end = row_start + row_count
    col_end = col_start + col_count
    for i_row in range(row_start, row_end):
        for i_col in range(col_start, col_end):
            print("{0}\t".format(t_ws.cell(i_row, i_col).value), end="")
        print()



def get_show_chassis_power_budget_statistics(contents_target_command, prompt_list):
    """
    show chassis power-budget-statisticsの実行結果を
    テーブル形式でリストに格納する
    """
    fpcs = []
    table = None
    for i in range(len(contents_target_command)):
        if commonLib.isPrompt(contents_target_command[i], prompt_list):
            if table != None and len(table)> 0:
                fpcs.append(table)
            continue
        elif "fpc" in contents_target_command[i] and "-------" in contents_target_command[i + 1]:
            if table != None and len(table)> 0:
                fpcs.append(table)
            table = []
            table.append(contents_target_command[i].replace("\n", ""))
            continue
        elif "-------" in contents_target_command[i]:
            continue
        elif " : " in contents_target_command[i]:
            flds = contents_target_command[i].split(":")
            for i in range(len(flds)):
                flds[i] = flds[i].strip()
            table.append(flds)
    return fpcs


def get_show_chassis_power_to_str(contents_target_command, prompt_list):
    """
    show chassis power* の実行結果を
    １つの文字列として連結する（空行は削除）
    """

    resStr = ""
    for i in range(len(contents_target_command)):
        if commonLib.isPrompt(contents_target_command[i], prompt_list):
            start = True
            continue
        if start and commonLib.isPrompt(contents_target_command[i], prompt_list):
            break
        else:
            if contents_target_command[i] == "\n":
                continue
            workStr = contents_target_command[i].replace("\n", "\r\n")
            resStr += workStr
    return resStr


def print_get_show_chassis_power_budget_statistics(fpcs):
    """テーブルへの格納結果を標準出力に出力する"""
    print("len(fpcs) = {0:d}".format(len(fpcs)))
    for table in fpcs:

        """ Output fpc Number """
        print(table[0])

        """ Output Column Title """
        for index in range(1, len(table)):
            print("\"{0}\",".format(table[index][0]), end="")
        print()

        """ Output Column Data """
        for index in range(1, len(table)):
            print("\"{0}\",".format(table[index][1]), end="")
        print()
    print()


def print_show_chassis_power_detail(res):
    """テーブルへの格納結果を標準出力に出力する"""
    print(res)


def output_show_chassis_power_budget_statistics(fpcs, t_ws, sPos):
    """テーブルへの格納結果をワークシートに出力する"""
    title_col = 4
    sPos[t_ws.title].i_col = 4
    for table in fpcs:
        '''
        """ Output fpc Number """
        if t_ws.cell(sPos[t_ws.title].i_row, title_col).value == None:
            # output cell
            t_ws.cell(sPos[t_ws.title].i_row, title_col).value = table[0]
            print("[{0:d}, {1:d}] = {2}".format(sPos[t_ws.title].i_row, title_col, table[0]))
        elif t_ws.cell(sPos[t_ws.title].i_row, title_col).value == table[0]:
            continue
        else:
            print("WorkSheets[{0}].Cell({1:d}, {2:d}) ... \"{3}\" != \"{4}\""
                .format(t_ws.title,  sPos[t_ws.title].i_row, title_col, t_ws.cell(sPos[t_ws.title].i_row, title_col).value, table[0]))
        '''
        """ Output Column Title """
        for index in range(1, len(table)):
            if t_ws.cell(1, title_col).value == None:
                # output cell
                t_ws.cell(1, title_col).value = table[index][0]
                print("[{0:d}, {1:d}] = {2}".format(1, title_col, table[index][0]))
            elif t_ws.cell(1, title_col).value == table[index][0]:
                continue
            else:
                print("WorkSheets[{0}].Cell({1:d}, {2:d}) ... \"{3}\" != \"{4}\""
                    .format(t_ws.title, 1, title_col, t_ws.cell(1, title_col).value, table[index][0]))
            title_col += 1

        """ Output Column Data """
        for index in range(1, len(table)):
            # output cell
            """
            t_ws.cell(sPos[t_ws.title].i_row + 2, sPos[t_ws.title].i_col).value = table[index][1]
            print("[{0:d}, {1:d}] = {2}".format(sPos[t_ws.title].i_row + 2, sPos[t_ws.title].i_col, table[index][1]))
            """
            t_ws.cell(sPos[t_ws.title].i_row, sPos[t_ws.title].i_col).value = table[index][1]
            print("[{0:d}, {1:d}] = {2}".format(sPos[t_ws.title].i_row, sPos[t_ws.title].i_col, table[index][1]))

            sPos[t_ws.title].i_col += 1

    # sPos[t_ws.title].i_row += 3
    sPos[t_ws.title].i_row += 1


def output_show_chassis_power_detail(resStr, t_ws, sPos):
    """格納結果をワークシートに出力する"""
    t_ws.cell(sPos[t_ws.title].i_row, sPos[t_ws.title].i_col).value = resStr
    sPos[t_ws.title].i_row += 1


if __name__ == "__main__":
    main()
